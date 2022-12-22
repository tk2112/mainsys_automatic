from modules.automatic.mainSysAuto import MainSysAuto
import os
import csv
import glob
import sqlite3
import re
import openpyxl
import datetime
import modules.const as const

const.EXCEL_COLUMN_ORDER_NO = 1
const.EXCEL_COLUMN_ORDER_NO_DETAIL = 2
const.EXCEL_COLUMN_ITEM_CODE = 3
const.EXCEL_COLUMN_CONTROL_NO_SUM = 4
const.EXCEL_COLUMN_AMOUNT = 5
const.EXCEL_COLUMN_SHIPMENT_PLAN = 6
const.EXCEL_COLUMN_NOTE = 7
const.EXCEL_COLUMN_DELIVERY = 8
const.EXCEL_ROW_COLUMNS = 1

class SvrDatabase(MainSysAuto):

    def storeOverseas(self):
        # 保存先外部svrのurl取得
        productDbFullPath = self.myWebDriver.jsonLoad['productDbFullPath']

        # DBファイルが存在しない場合、処理を抜ける
        if not os.path.isfile(productDbFullPath):
            return

        # 出荷情報取得
        fileList = glob.glob(self.myWebDriver.downloadDir + "/出荷情報*.csv")

        # 正規表現パターン：海外営業拠点コード
        patternOrderNo = r'J\d{9}-070-(060|061)'

        # 正規表現パターン：日付
        patternDate = r'^(\d{4})/(\d{1,2})/(\d{1,2})$'

        # 海外営業搬入予定一覧のEXCEL作成
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = 'test_sheet_1'

        for file in fileList:
            with open(file, 'r') as fOrigin:

                reader = csv.reader(fOrigin)
                header = next(reader)
                sheet.cell(row=const.EXCEL_ROW_COLUMNS,column=const.EXCEL_COLUMN_ORDER_NO).value = '受注番号'
                sheet.cell(row=const.EXCEL_ROW_COLUMNS,column=const.EXCEL_COLUMN_ORDER_NO_DETAIL).value = '受注明細番号'
                sheet.cell(row=const.EXCEL_ROW_COLUMNS,column=const.EXCEL_COLUMN_ITEM_CODE).value = '品目コード'
                sheet.cell(row=const.EXCEL_ROW_COLUMNS,column=const.EXCEL_COLUMN_CONTROL_NO_SUM).value = '管理番号'
                sheet.cell(row=const.EXCEL_ROW_COLUMNS,column=const.EXCEL_COLUMN_AMOUNT).value = '数量'
                sheet.cell(row=const.EXCEL_ROW_COLUMNS,column=const.EXCEL_COLUMN_SHIPMENT_PLAN).value = '出荷予定日'
                sheet.cell(row=const.EXCEL_ROW_COLUMNS,column=const.EXCEL_COLUMN_NOTE).value = '摘要'
                sheet.cell(row=const.EXCEL_ROW_COLUMNS,column=const.EXCEL_COLUMN_DELIVERY).value = '納品先'
                currentExcelRow = const.EXCEL_ROW_COLUMNS
                keyDics = {}
                
                for row in reader:
                    currentOrderNo = row[0]
                    currentOrderNoDetail = row[1]
                    currentKey = currentOrderNo + '_' + currentOrderNoDetail
                    
                    if not re.search(patternOrderNo ,currentOrderNo):
                        pass
                    elif currentKey in keyDics:
                        pass
                    else:
                        currentExcelRow += 1

                        sheet.cell(row=currentExcelRow,column=const.EXCEL_COLUMN_ORDER_NO).number_format = '@'
                        sheet.cell(row=currentExcelRow,column=const.EXCEL_COLUMN_ORDER_NO_DETAIL).number_format = '0'
                        sheet.cell(row=currentExcelRow,column=const.EXCEL_COLUMN_ITEM_CODE).number_format = '@'
                        sheet.cell(row=currentExcelRow,column=const.EXCEL_COLUMN_CONTROL_NO_SUM).number_format = '@'
                        sheet.cell(row=currentExcelRow,column=const.EXCEL_COLUMN_AMOUNT).number_format = '0'
                        sheet.cell(row=currentExcelRow,column=const.EXCEL_COLUMN_SHIPMENT_PLAN).number_format = 'yyyy/mm/dd'
                        sheet.cell(row=currentExcelRow,column=const.EXCEL_COLUMN_NOTE).number_format = '@'
                        sheet.cell(row=currentExcelRow,column=const.EXCEL_COLUMN_DELIVERY).number_format = '@'

                        matchDate = re.search(patternDate ,row[18])

                        sheet.cell(row=currentExcelRow,column=const.EXCEL_COLUMN_ORDER_NO).value = currentOrderNo
                        sheet.cell(row=currentExcelRow,column=const.EXCEL_COLUMN_ORDER_NO_DETAIL).value = int(currentOrderNoDetail)
                        sheet.cell(row=currentExcelRow,column=const.EXCEL_COLUMN_ITEM_CODE).value = row[16]
                        sheet.cell(row=currentExcelRow,column=const.EXCEL_COLUMN_CONTROL_NO_SUM).value = row[132]
                        sheet.cell(row=currentExcelRow,column=const.EXCEL_COLUMN_AMOUNT).value = int(row[133])
                        sheet.cell(row=currentExcelRow,column=const.EXCEL_COLUMN_NOTE).value = row[93]
                        sheet.cell(row=currentExcelRow,column=const.EXCEL_COLUMN_DELIVERY).value = row[2]

                        if matchDate:
                            currentYear = int(matchDate[1])
                            currentMonth = int(matchDate[2])
                            currentDay = int(matchDate[3])

                            sheet.cell(row=currentExcelRow,column=const.EXCEL_COLUMN_SHIPMENT_PLAN).value = datetime.datetime(currentYear, currentMonth, currentDay)

                        keyDics[currentKey] = 1

        # 列幅自動調整
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter

            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))

            adjusted_width = (max_length + 2) * 1.2
            
            sheet.column_dimensions[column].width = adjusted_width

        wb.save(self.myWebDriver.currentDir + '/test.xlsx')
        
        # con = sqlite3.connect(productDbFullPath)
        # cur = con.cursor()

        # sql = 'select * from mt_classifications;'
        # cur.execute(sql)
 
        # for row in cur:
        #     print(row[0], row[1])

        # con.close()